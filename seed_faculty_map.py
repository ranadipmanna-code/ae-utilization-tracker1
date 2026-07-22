"""
Rebuild `core_ae_faculty_map` from the authoritative source:
    AE_Alignment-PAN_India.xlsx  ->  sheet "AE-Trainer Alignment (Updated)"

Mapping used:
    Trainer email = col "Trainer's Email id"
    Core AE email = col "AE Email ID"   (the AE SPOC for that trainer)

This makes the app's session counts match CMIS for each Core AE, because the
faculty list now comes straight from the official alignment.

The table stores up to 5 faculty per row (faculty_1..faculty_5); a Core AE with
more than 5 trainers spans multiple rows, which the app already handles.

Run:
  docker run --rm -v "${PWD}:/app" -w /app python:3.11-slim bash -c \
    "pip install -q -r requirements.txt openpyxl && python seed_faculty_map.py"
"""
import sys
import tomllib
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote_plus

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

SECRETS = Path(".streamlit/secrets.toml")
XLSX = Path("AE_Alignment-PAN_India.xlsx")
SHEET = "AE-Trainer Alignment (Updated)"

TRAINER_COL = 2   # "Trainer's Email id"
COREAE_COL = 10   # "AE Email ID"
FACULTY_PER_ROW = 5

NA = {"", "#n/a", "n/a", "none", "null"}

# The alignment file lists some Extended AEs in the trainer column (they report
# to a Core AE too). Those must NOT become "faculty". Exclude the known
# Extended AEs so core_ae_faculty_map holds real trainers only.
EXTENDED_AES = {
    "priyanka.roy@anudip.org", "shanmukh.adala@anudip.org", "kundan.sinha@anudip.org",
    "sabreena.ramzan@anudip.org", "divya.ns@anudip.org", "pallav.punit@anudip.org",
    "anirudhha.sharma@anudip.org", "grk.mahalakshmi@anudip.org", "dipankar.biswas@anudip.org",
    "aarti.kumari@anudip.org", "pranjya.das@anudip.org", "pulak.bhattacharya@anudip.org",
    "priyanka.nongkhlaw@anudip.org",
}


def eng():
    cfg = tomllib.load(open(SECRETS, "rb"))["appdb"]
    url = (f"mysql+pymysql://{quote_plus(str(cfg['user']))}:{quote_plus(str(cfg['password']))}"
           f"@{cfg['host']}:{cfg['port']}/{cfg['database']}?charset=utf8mb4")
    return create_engine(url)


def main():
    if not SECRETS.exists():
        sys.exit("❌ .streamlit/secrets.toml not found.")
    if not XLSX.exists():
        sys.exit(f"❌ {XLSX} not found — put the alignment file in this folder.")

    ws = load_workbook(XLSX, data_only=True)[SHEET]

    core_to_trainers = defaultdict(list)
    unmapped = []
    for r in range(2, ws.max_row + 1):
        te = ws.cell(r, TRAINER_COL).value
        core = ws.cell(r, COREAE_COL).value
        if not te:
            continue
        te = str(te).strip().lower()
        # skip Extended AEs — they aren't faculty even if listed as trainers
        if te in EXTENDED_AES:
            continue
        core_norm = str(core).strip().lower() if core else ""
        if core_norm in NA:
            unmapped.append(te)
            continue
        # de-dupe within a Core AE
        if te not in core_to_trainers[core_norm]:
            core_to_trainers[core_norm].append(te)

    e = eng()
    with e.begin() as c:
        c.execute(text("DELETE FROM core_ae_faculty_map"))
        rows = 0
        for core, trainers in sorted(core_to_trainers.items()):
            for i in range(0, len(trainers), FACULTY_PER_ROW):
                chunk = trainers[i:i + FACULTY_PER_ROW] + [None] * FACULTY_PER_ROW
                c.execute(
                    text(
                        "INSERT INTO core_ae_faculty_map "
                        "(core_ae_email, faculty_1, faculty_2, faculty_3, faculty_4, faculty_5) "
                        "VALUES (:c,:f1,:f2,:f3,:f4,:f5)"
                    ),
                    {"c": core, "f1": chunk[0], "f2": chunk[1], "f3": chunk[2],
                     "f4": chunk[3], "f5": chunk[4]},
                )
                rows += 1

    total_trainers = sum(len(v) for v in core_to_trainers.values())
    print("✅ core_ae_faculty_map rebuilt from AE_Alignment-PAN_India.xlsx")
    print(f"   Core AEs   : {len(core_to_trainers)}")
    print(f"   Trainers   : {total_trainers} mapped")
    print(f"   Map rows   : {rows}")
    print(f"   ⚠️  Unmapped: {len(unmapped)} trainers have no AE (col 'AE Email ID' = #N/A)")
    print()
    print("   Trainers per Core AE:")
    for core in sorted(core_to_trainers, key=lambda k: -len(core_to_trainers[k])):
        print(f"     {core:38s} {len(core_to_trainers[core])}")
    if unmapped:
        print()
        print("   These trainers won't appear under any Core AE until assigned in the")
        print("   alignment file (first 10):")
        for t in unmapped[:10]:
            print(f"     {t}")


if __name__ == "__main__":
    main()
