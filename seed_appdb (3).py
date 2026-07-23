"""
One-time seed for the App DB (Anudip_AE_Team).

Populates:
  * user_roles          — admin + 12 Core AE + 13 Extended AE logins
  * core_ae_faculty_map — each Core AE with up to 5 faculty emails,
                          taken from the real AE_Alignment Excel (trainers whose
                          AE SPOC is that Core AE, chunked into rows of 5).

Run once after the tables exist:

    python seed_appdb.py

Idempotent-ish: it clears user_roles and core_ae_faculty_map first, then
re-inserts, so re-running gives a clean roster. It NEVER touches CMIS.
"""
import sys
import tomllib
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote_plus

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

SECRETS = Path(".streamlit/secrets.toml")
EXCEL = Path("ae_alignment.xlsx")

# 12 canonical Core AEs (name -> email); mirrors the FastAPI seed roster.
CORE_AE = {
    "rashmi.mukherjee@anudip.org": "Rashmi Mukherjee",
    "tanmoy.bose@anudip.org": "Tanmoy Bose",
    "milan.biswas@anudip.org": "Milan Biswas",
    "biswajit.chakraborty@anudip.org": "Biswajit Chakraborty",
    "brahma@anudip.org": "Bramha Ji",
    "sapna.yadav@anudip.org": "Sapna Yadav",
    "karishma.tiwari@anudip.org": "Karishma Tiwari",
    "navamita.talukdar@anudip.org": "Navamita",
    "susmita.chakrabarty@anudip.org": "Susmita Chakraborty",
    "sirivennela.gaddam@anudip.org": "Siri",
    "arnab.roy@anudip.org": "Arnab",
    "madhu.soni@anudip.org": "Madhu Soni",
}

# 13 Extended AEs (email -> name).
#
# IMPORTANT: these addresses are the ones CMIS actually stores in
# upcoming_trainer_utilization_view.email_id, NOT the "tidy" firstname.lastname
# form. The app joins CMIS on an exact email string match, so any drift here
# means that Extended AE silently shows an empty Calendar/Sessions tab even
# though their sessions exist. Six were wrong before and are marked below.
# Verified against CMIS on 2026-07-23; run `python seed_appdb.py --verify`
# to re-check before seeding.
EXTENDED_AE = {
    "priyanka_roy@anudip.org": "Priyanka Roy",            # was priyanka.roy@   (underscore)
    "kundan.sinha@anudip.org": "Kundan Sinha",
    "sabreena.ramzan@anudip.org": "Sabreena Ramzan",
    "divya.s@anudip.org": "Divya NS",                     # was divya.ns@
    "pallav.punit@anudip.org": "Pallav Punit",
    "aniruddha.sharma@anudip.org": "Anirudhha Sharma",    # was anirudhha.sharma@
    "mahalakshmi.grk@anudip.org": "GRK Mahalakshmi",      # was grk.mahalakshmi@ (tokens reversed)
    "dipankar.biswas@anudip.org": "Dipankar Biswas",
    "aarti.kumari@anudip.org": "Aarti Kumari",
    "pranjya.dash@anudip.org": "Pranjya Priyadarsani Das",  # was pranjya.das@
    "pulak@anudip.org": "Pulak Bhattacharya",             # was pulak.bhattacharya@
    "priyanka.nongkhlaw@anudip.org": "Priyanka Nongkhlaw",
    "shanmukh.addala@anudip.org": "Shanmukh Adala",       # was shanmukh.adala@
}

# Core AEs are deliberately NOT resolved against CMIS. They observe rather than
# deliver, so they have no row in the trainer view at all — and attempting a
# match is actively dangerous: 'biswajit.chakraborty' fuzzy-matches the real
# but unrelated trainer 'ajit.chakraborty' (126 sessions) at 0.88 similarity.

CORE_EMAILS = set(CORE_AE)
SPOC_NAME_OVERRIDE = {"Arnab Roy (IBM)": "arnab.roy@anudip.org"}

C_TRAINER_EMAIL, C_AE_EMAIL, C_SPOC_NAME = 1, 9, 6


def load_secrets():
    with open(SECRETS, "rb") as f:
        return tomllib.load(f)


def engine():
    cfg = load_secrets()["appdb"]
    pwd = quote_plus(str(cfg["password"]))
    user = quote_plus(str(cfg["user"]))
    url = (f"mysql+pymysql://{user}:{pwd}"
           f"@{cfg['host']}:{cfg['port']}/{cfg['database']}?charset=utf8mb4")
    return create_engine(url)


def _clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("#N/A", "N/A", "None", "NA") else s


def build_faculty_map():
    """core_ae_email -> [faculty emails] from the Excel."""
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb.active
    m = defaultdict(list)
    seen = set()
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 16)]
        temail = _clean(v[C_TRAINER_EMAIL]).lower()
        if not temail or temail in seen or temail in CORE_EMAILS or temail in EXTENDED_AE:
            continue
        ce = _clean(v[C_AE_EMAIL]).lower()
        if ce not in CORE_EMAILS:
            ce = SPOC_NAME_OVERRIDE.get(_clean(v[C_SPOC_NAME]))
        if not ce:
            continue
        seen.add(temail)
        m[ce].append(temail)
    return m


def verify_against_cmis() -> int:
    """Check every Extended AE email really exists in CMIS. Returns bad count.

    Cheap insurance: the whole class of bug this seed had was addresses that
    look plausible but match nothing in CMIS, which fails silently at runtime
    instead of loudly here.
    """
    cfg = load_secrets()["cmis"]
    url = (f"mysql+pymysql://{quote_plus(str(cfg['user']))}:"
           f"{quote_plus(str(cfg['password']))}"
           f"@{cfg['host']}:{cfg['port']}/{cfg['database']}?charset=utf8mb4")
    eng = create_engine(url)
    with eng.connect() as conn:
        known = {
            str(r[0]).strip().lower()
            for r in conn.execute(text(
                "SELECT DISTINCT email_id FROM upcoming_trainer_utilization_view "
                "WHERE email_id IS NOT NULL"
            ))
        }
    bad = 0
    print(f"CMIS knows {len(known)} trainer emails.\n")
    for email, name in EXTENDED_AE.items():
        ok = email.lower() in known
        print(f"  {'OK ' if ok else 'MISSING'}  {email:34} {name}")
        bad += (not ok)
    if bad:
        print(f"\n{bad} Extended AE email(s) are not in CMIS — fix before seeding.")
    else:
        print("\nAll Extended AE emails resolve in CMIS.")
    return bad


def main():
    if "--verify" in sys.argv:
        sys.exit(1 if verify_against_cmis() else 0)
    if not SECRETS.exists():
        sys.exit("❌ .streamlit/secrets.toml not found.")
    if not EXCEL.exists():
        sys.exit("❌ ae_alignment.xlsx not found next to this script.")

    fac_map = build_faculty_map()
    eng = engine()

    # The real user_roles table has a NOT NULL `password` column. Use the
    # shared demo password from secrets so logins work.
    shared_pwd = load_secrets().get("auth", {}).get("shared_password", "Password123!")

    with eng.begin() as conn:
        # --- user_roles ---
        conn.execute(text("DELETE FROM user_roles"))
        conn.execute(
            text("INSERT INTO user_roles (email, name, role, password) "
                 "VALUES ('admin1@anudip.org','Admin One','admin',:p)"),
            {"p": shared_pwd},
        )
        for email, name in CORE_AE.items():
            conn.execute(
                text("INSERT INTO user_roles (email, name, role, password) VALUES (:e,:n,'core_ae',:p)"),
                {"e": email, "n": name, "p": shared_pwd},
            )
        for email, name in EXTENDED_AE.items():
            conn.execute(
                text("INSERT INTO user_roles (email, name, role, password) VALUES (:e,:n,'extended_ae',:p)"),
                {"e": email, "n": name, "p": shared_pwd},
            )

        # --- core_ae_faculty_map: chunk faculty into rows of up to 5 ---
        conn.execute(text("DELETE FROM core_ae_faculty_map"))
        total_rows = 0
        for ce, facs in fac_map.items():
            for i in range(0, len(facs), 5):
                chunk = facs[i:i + 5] + [None] * 5
                conn.execute(
                    text(
                        "INSERT INTO core_ae_faculty_map "
                        "(core_ae_email, faculty_1, faculty_2, faculty_3, faculty_4, faculty_5) "
                        "VALUES (:c,:f1,:f2,:f3,:f4,:f5)"
                    ),
                    {"c": ce, "f1": chunk[0], "f2": chunk[1], "f3": chunk[2], "f4": chunk[3], "f5": chunk[4]},
                )
                total_rows += 1

    print("✅ Seed complete.")
    print(f"   user_roles: 1 admin + {len(CORE_AE)} core + {len(EXTENDED_AE)} extended")
    print(f"   core_ae_faculty_map: {total_rows} rows across {len(fac_map)} Core AEs")
    print("   Login password is whatever you set in secrets.toml [auth].shared_password")


if __name__ == "__main__":
    main()
